import os
import pandas as pd
from pydantic import BaseModel, Field
from typing import Dict, Any, Tuple, Optional, List, AsyncGenerator
from groq import Groq
from openai import OpenAI
from google import genai
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelTools(BaseModel):
    """
    Handles all CRUD operations on Excel files using a priority-based 
    loading system (modified vs original)[cite: 4].
    """
    source_dir: str = Field(default="data_original", description="Folder for raw data")
    target_dir: str = Field(default="data_modified", description="Folder for saved changes")
    
    files: Dict[str, str] = Field(default_factory=lambda: {
    "properties": "Real Estate Listings.xlsx",
    "marketing": "Marketing Campaigns.xlsx"
    })

    def _get_path(self, file_key: str, folder: str) -> str:
        """Helper to build absolute file paths."""
        return os.path.join(folder, self.files[file_key])

    def _load_data(self, file_key: str) -> Tuple[Optional[pd.DataFrame], Optional[str]]:
        if file_key not in self.files:
            return None, f"Invalid key: {file_key}. Use 'properties' or 'marketing'."

        mod_path = self._get_path(file_key, self.target_dir)
        orig_path = self._get_path(file_key, self.source_dir)
        path = mod_path if os.path.exists(mod_path) else orig_path

        if not os.path.exists(path):
            return None, f"File {self.files[file_key]} not found in source or target folders."
            
        try:
            df = pd.read_excel(path)
            currency_cols = [c for c in df.columns if any(x in c for x in
                            ["Price", "Budget", "Spent", "Revenue", "Amount"])]
            for col in currency_cols:
                if df[col].dtype == object:
                    df[col] = pd.to_numeric(
                        df[col].astype(str).str.replace(r'[\$,]', '', regex=True),
                        errors='coerce'
                    )
            int_cols = [c for c in df.columns if any(x in c for x in
                    ["Impressions", "Clicks", "Conversions", "Bedrooms",
                        "Bathrooms", "Footage", "Year"])]
            for col in int_cols:
                if df[col].dtype == object:
                    df[col] = pd.to_numeric(
                        df[col].astype(str).str.replace(',', '', regex=False),
                        errors='coerce'
                    )

            return df, None
        except Exception as e:
            return None, f"Read Error: {str(e)}"
        
    def _save_data(self, file_key: str, df: pd.DataFrame) -> str:
        if not os.path.exists(self.target_dir):
            os.makedirs(self.target_dir)

        path = self._get_path(file_key, self.target_dir)
        try:
            date_cols = [c for c in df.columns if "date" in c.lower()]
            for col in date_cols:
                df[col] = pd.to_datetime(df[col]).dt.date

            currency_cols = [c for c in df.columns if any(x in c for x in 
                            ["Price", "Budget", "Spent", "Revenue", "Amount"])]
            for col in currency_cols:
                if df[col].dtype == object:
                    df[col] = df[col].astype(str).str.replace(r'[\$,]', '', regex=True)
                    df[col] = pd.to_numeric(df[col], errors='coerce')

            int_cols = [c for c in df.columns if any(x in c for x in 
                    ["Impressions", "Clicks", "Conversions", "Bedrooms", 
                        "Bathrooms", "Footage", "Year"])]
            for col in int_cols:
                if df[col].dtype == object:
                    df[col] = df[col].astype(str).str.replace(',', '')
                    df[col] = pd.to_numeric(df[col], errors='coerce')

            wb = openpyxl.Workbook()
            ws = wb.active

            header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            header_align = Alignment(horizontal="center", vertical="center")

            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align

            for sheet_row, (row_idx, row) in enumerate(df.iterrows(), start=1):
                for col_idx, (col_name, value) in enumerate(row.items(), 1):
                    cell = ws.cell(row=sheet_row + 1, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="center")

                    if col_name in date_cols:
                        cell.number_format = 'MM/DD/YYYY'

                    elif col_name in currency_cols:
                        cell.number_format = '"$"#,##0.00'

                    elif col_name in int_cols:
                        cell.number_format = '#,##0'

            for col_idx, col_name in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_idx)
                max_len = max(
                    len(str(col_name)),
                    df.iloc[:, col_idx - 1].astype(str).str.len().max()
                )
                ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

            wb.save(path)
            return f"Success: Modified data saved to {path}."
        except Exception as e:
            return f"Write Error: {str(e)}"

    def get_schema(self, file_key: str) -> str:
        """Returns column names and types so the LLM knows the data structure[cite: 4]."""
        df, err = self._load_data(file_key)
        if err: return err
        return f"Columns for {file_key}: {df.dtypes.to_dict()}"
    
    def get_row_index(self, file_key: str, query_str: str) -> str:
        """Returns the DataFrame index of rows matching a query — use before delete_record or update_record."""
        df, err = self._load_data(file_key)
        if err: return err
        try:
            clean_query = query_str
            # Only strip outer quotes if entire query is wrapped
            if (query_str.startswith("'") and query_str.endswith("'")) or \
            (query_str.startswith('"') and query_str.endswith('"')):
                clean_query = query_str[1:-1]

            # Normalize curly quotes
            clean_query = clean_query.replace('\u201c', '"').replace('\u201d', '"')
            clean_query = clean_query.replace('\u2018', "'").replace('\u2019', "'")

            for col in df.columns:
                if " " in col and col in clean_query and f"`{col}`" not in clean_query:
                    clean_query = clean_query.replace(col, f"`{col}`")

            result = df.query(clean_query)
            if result.empty:
                return "No matching rows found."
            return f"Matching row indices: {result.index.tolist()}"
        except Exception as e:
            return f"Query Error: {str(e)}"

        
    def delete_missing_rows(self, file_key: str) -> str:
        """Deletes all rows that have any missing/NaN values."""
        df, err = self._load_data(file_key)
        if err: return err
        try:
            before = len(df)
            df = df.dropna()
            after = len(df)
            removed = before - after
            if removed == 0:
                return "No missing values found. Nothing was deleted."
            return self._save_data(file_key, df) + f" Removed {removed} row(s) with missing values."
        except Exception as e:
            return f"Delete Error: {str(e)}"
        
    def query_data(self, file_key: str, query_str: str) -> str:
        df, err = self._load_data(file_key)
        if err: return err
        try:
            clean_query = query_str
            if (query_str.startswith("'") and query_str.endswith("'")) or \
            (query_str.startswith('"') and query_str.endswith('"')):
                clean_query = query_str[1:-1]

            clean_query = clean_query.replace('\u201c', '"').replace('\u201d', '"')
            clean_query = clean_query.replace('\u2018', "'").replace('\u2019', "'")

            for col in df.columns:
                if " " in col and col in clean_query and f"`{col}`" not in clean_query:
                    clean_query = clean_query.replace(col, f"`{col}`")

            result = df.query(clean_query)
            if result.empty:
                return "No records found matching query."

            currency_cols = [c for c in result.columns if any(x in c for x in
                            ["Price", "Budget", "Spent", "Revenue", "Amount"])]
            for col in currency_cols:
                result = result.copy()
                result[col] = result[col].apply(lambda x: f"${x:,.2f}" if pd.notnull(x) else "")

            return result.to_markdown(index=True)
        except Exception as e:
            return f"Query Error: {str(e)}. Please check your syntax."
    
    def compute_and_query(self, file_key: str, new_column: str, formula: str, top_n: int = None, ascending: bool = True, query_str: str = None) -> str:
        """
        Adds a computed column and optionally filters/sorts results.
        formula: pandas expression e.g. '`Revenue Generated` / `Budget Allocated`'
        top_n: return only top N rows
        ascending: True for worst-first, False for best-first
        query_str: optional filter after computing
        """
        df, err = self._load_data(file_key)
        if err: return err
        try:
            df[new_column] = df.eval(formula)

            if query_str:
                clean_query = query_str
                for col in df.columns:
                    if " " in col and col in clean_query and f"`{col}`" not in clean_query:
                        clean_query = clean_query.replace(col, f"`{col}`")
                df = df.query(clean_query)

            if top_n:
                df = df.nsmallest(top_n, new_column) if ascending else df.nlargest(top_n, new_column)

            result = df.to_markdown(index=True)
            return f"NOTE: The first column is the DataFrame row index — use these numbers directly in delete_record or update_record.\n\n{result}"
        except Exception as e:
            return f"Compute Error: {str(e)}"

    def add_record(self, file_key: str, data: Dict[str, Any]) -> str:
        df, err = self._load_data(file_key)
        if err: return err
        try:
            # Strip backticks from keys
            clean_data = {k.strip('`'): v for k, v in data.items()}
            
            # Validate columns
            invalid = [k for k in clean_data if k not in df.columns]
            if invalid:
                return f"Insert Error: Unknown columns {invalid}. Available: {list(df.columns)}"
            
            new_row = pd.DataFrame([clean_data])
            df = pd.concat([df, new_row], ignore_index=True)
            return self._save_data(file_key, df)
        except Exception as e:
            return f"Insert Error: {str(e)}"

    def add_records(self, file_key: str, records: List[Dict[str, Any]]) -> str:
        """Inserts multiple rows in a single operation."""
        df, err = self._load_data(file_key)
        if err: return err
        try:
            clean_records = [{k.strip('`'): v for k, v in row.items()} for row in records]
            for record in clean_records:
                invalid = [k for k in record if k not in df.columns]
                if invalid:
                    return f"Insert Error: Unknown columns {invalid}. Available: {list(df.columns)}"
            new_rows = pd.DataFrame(clean_records)
            df = pd.concat([df, new_rows], ignore_index=True)
            return self._save_data(file_key, df) + f" Added {len(clean_records)} row(s)."
        except Exception as e:
            return f"Insert Error: {str(e)}"

    def update_record(self, file_key: str, row_index: int, updates: Dict[str, Any]) -> str:
        df, err = self._load_data(file_key)
        if err: return err
        try:
            # Strip backticks from keys in case LLM passes them
            clean_updates = {k.strip('`'): v for k, v in updates.items()}
            
            for key, value in clean_updates.items():
                if key not in df.columns:
                    return f"Update Error: Column '{key}' does not exist. Available columns: {list(df.columns)}"
                df.at[int(row_index), key] = value
                
            return self._save_data(file_key, df)
        except Exception as e:
            return f"Update Error: {str(e)}"

    def delete_record(self, file_key: str, row_index: int) -> str:
        """Removes a record by index[cite: 4]."""
        df, err = self._load_data(file_key)
        if err: return err
        try:
            df = df.drop(index=int(row_index))
            return self._save_data(file_key, df)
        except Exception as e:
            return f"Delete Error: {str(e)}"
        
    def find_missing_rows(self, file_key: str) -> str:
        """Shows all rows that have any missing/NaN values."""
        df, err = self._load_data(file_key)
        if err: return err
        try:
            result = df[df.isnull().any(axis=1)]
            if result.empty:
                return "No missing values found."
            return f"Found {len(result)} row(s) with missing values:\n{result.to_markdown(index=True)}"
        except Exception as e:
            return f"Error: {str(e)}"
    
    def summarize_data(self, file_key: str, group_by: str = None, column: str = None) -> str:
        """
        Returns summary statistics. 
        - No args: describes all numeric columns
        - group_by + column: returns mean/sum/max/min per group
        """
        df, err = self._load_data(file_key)
        if err: return err
        try:
            if group_by and column:
                group_col = group_by.strip('`')
                agg_col = column.strip('`')
                result = df.groupby(group_col)[agg_col].agg(['mean', 'sum', 'max', 'min', 'count']).reset_index()
                result.columns = [group_col, 'Mean', 'Total', 'Max', 'Min', 'Count']
                result = result.sort_values('Mean', ascending=False)
                return result.to_markdown(index=False)
            else:
                return df.describe().to_markdown()
        except Exception as e:
            return f"Summary Error: {str(e)}"
    
    def delete_records(self, file_key: str, row_indices: list) -> str:
        """Removes multiple records by index in a single operation."""
        df, err = self._load_data(file_key)
        if err: return err
        try:
            valid_indices = [int(i) for i in row_indices if int(i) in df.index]
            if not valid_indices:
                return "No valid indices found to delete."
            df = df.drop(index=valid_indices)
            return self._save_data(file_key, df) + f" Removed {len(valid_indices)} row(s)."
        except Exception as e:
            return f"Delete Error: {str(e)}"


# --- Agent Core: Reasoning & LLM Routing [cite: 13, 15] ---

SYSTEM_PROMPT = """
You are a data assistant that queries Excel files using tools.

STRICT RULES:
1. Only call tools when the user asks a data-related question.
   For greetings like 'hi', respond with a Final Answer immediately.
2. Output ONLY one Thought + Action pair per turn, then STOP.
   NEVER write Observation yourself — the system provides it.
3. Column names with spaces MUST use backticks in query_str.
   Example: query_data('properties', "`List Price` > 500000")
4. Always call get_schema first if you don't know the column names.
5. ALWAYS use positional arguments in tool calls, never keyword arguments.
   CORRECT:   add_records('properties', [{'Listing ID': 'X', 'City': 'Y'}])
   INCORRECT: add_records(file_key='properties', records=[{'Listing ID': 'X'}])

FORMAT:
Thought: <your reasoning>
Action: tool_name(args)

OR:

Final Answer: <your response to the user>

Available Tools:
- get_schema(file_key): Get columns for 'properties' or 'marketing'.

- query_data(file_key, query_str): Filter rows using pandas query syntax.

- add_record(file_key, data): Insert a new row as a dictionary.

- add_records(file_key, records): Insert multiple rows at once as a list of dictionaries.
  Use this instead of calling add_record multiple times.
  Example: add_records('properties', [{'Listing ID': 'DUMMY-001', 'City': 'Austin', ...}, {'Listing ID': 'DUMMY-002', ...}])
  ALWAYS prefer this over multiple add_record calls when inserting more than one row.

- update_record(file_key, row_index, updates): Update a row by index.

- delete_record(file_key, row_index): Delete a row by index.

- get_row_index(file_key, query_str): Get the exact row index before calling delete_record or update_record.
  ALWAYS use this before deleting or updating. Example workflow:
  1. get_row_index('marketing', "`Revenue Generated` == 30")  → returns index [102]
  2. delete_record('marketing', 102)
  
- delete_missing_rows(file_key): Delete all rows that contain any missing or NaN values.
  Use this instead of query_data when the user asks to remove incomplete/missing rows.
  
- find_missing_rows(file_key): Preview all rows with any missing/NaN values.
  Use this when the user asks to check for or show missing/incomplete data.
  Always call this before delete_missing_rows so the user can confirm what will be deleted.
  
- summarize_data(file_key, group_by, column): Use for averages, totals, comparisons across groups.
  Example: summarize_data('marketing', 'Channel', 'Revenue Generated')
  Use this instead of query_data when the user asks "which X has highest/average/total Y".
  
- compute_and_query(file_key, new_column, formula, top_n, ascending, query_str):
  Use when the user asks about calculated metrics like ROI, profit margin, or any derived value.
  formula uses pandas eval syntax e.g. '`Revenue Generated` / `Budget Allocated`'
  Example for top 5 List Price:  compute_and_query('properties', 'Rank', '`List Price`', top_n=5, ascending=False)
  Example for 3 worst ROI:       compute_and_query('marketing', 'ROI', '`Revenue Generated` / `Budget Allocated`', top_n=3, ascending=True)
  Example for 3 best ROI:        compute_and_query('marketing', 'ROI', '`Revenue Generated` / `Budget Allocated`', top_n=3, ascending=False)
  ALWAYS use keyword arguments for top_n, ascending, and query_str to avoid conflicts.
  Use this instead of query_data for any question involving division, multiplication, or derived columns.
  IMPORTANT: The first column of the result is the exact DataFrame row index. Use these indices directly
  in delete_record, delete_records, or update_record. Do NOT call get_row_index afterward.

- delete_records(file_key, row_indices): Delete multiple records at once by passing a list of indices.
  Use this instead of calling delete_record multiple times.
  Example: delete_records('properties', [311, 403, 469, 503, 802])
  ALWAYS prefer this over multiple delete_record calls when deleting more than one row.
  
Additional Thoughts:
- For "is X proportional to Y across groups" questions:
  1. Use summarize_data to get totals for both X and Y per group
  2. Reason over both tables yourself to assess proportionality
  3. Do NOT use compute_and_query with top_n=1 for comparison questions
  
File keys: 'properties' = Real Estate Listings, 'marketing' = Marketing Campaigns.
"""

class AIAgent:
    def __init__(self, model_choice: str, api_keys: dict):
        self.model_choice = str(model_choice) if model_choice else "Groq (Llama 3.3)"
        self.tools = ExcelTools()
        self.history = []
        self._pending_operation = None  # stores (func, args) waiting for confirmation
        
        self.groq_client = Groq(api_key=api_keys.get("groq"))
        self.nv_client = OpenAI(
            base_url="https://integrate.api.nvidia.com/v1",
            api_key=api_keys.get("nvidia")
        )
        self.gemini_client = genai.Client(api_key=api_keys.get("gemini"))
        
    def _call_llm(self, prompt: str) -> str:
        try:
            choice = self.model_choice or "Groq (Llama 3.3)"
            
            if "Groq" in choice:
                res = self.groq_client.chat.completions.create(
                    model="llama-3.3-70b-versatile",
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.2,
                    max_tokens=1000
                )
                return res.choices[0].message.content or ""
            
            elif "NVIDIA" in choice:
                res = self.nv_client.chat.completions.create(
                    model="meta/llama-3.3-70b-instruct",
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.2,
                    max_tokens=1000
                )
                return res.choices[0].message.content or ""
            
            else:  
                res = self.gemini_client.models.generate_content(
                    model="gemini-2.5-flash-lite",
                    contents=prompt
                )
                return res.text or ""
        except Exception as e:
            return f"Critical API Error: {str(e)}"
        
    def confirm_pending(self):
        """Execute the pending operation after user confirms."""
        if self._pending_operation:
            func, args, kwargs = self._pending_operation
            result = func(*args, **kwargs)
            self._pending_operation = None
            self.history.append({"role": "system", "content": "Operation completed successfully. If the original task is fully done, respond with a Final Answer summarizing what was accomplished."})
            return result
        return None

    def cancel_pending(self):
        """Cancel the pending operation."""
        self._pending_operation = None
        self._last_observation = "Operation cancelled by user."

    async def continue_after_confirm(self) -> AsyncGenerator[str, None]:
        """Resume the ReAct loop after user confirms a write operation."""
        for _ in range(5):
            full_prompt = SYSTEM_PROMPT + "\n" + "\n".join(
                [f"{m['role']}: {m['content']}" for m in self.history]
            )
            response = self._call_llm(full_prompt)
            if not response:
                yield "No response from model."
                break

            lines = response.split("\n")
            truncated_lines = []
            for line in lines:
                truncated_lines.append(line)
                if line.strip().startswith("Action:") and "(" in line:
                    break
            response = "\n".join(truncated_lines)

            yield response

            if "Final Answer:" in response:
                self.history.append({"role": "assistant", "content": response})
                break

            elif "Action:" in response:
                action_line = next(
                    (line for line in response.split("\n") if line.strip().startswith("Action:") and "(" in line),
                    None
                )
                if not action_line or "none" in action_line.lower():
                    self.history.append({"role": "assistant", "content": response})
                    break

                obs = self._execute_tool(action_line)
                self.history.append({"role": "assistant", "content": response})

                if obs.startswith("__CONFIRM__:"):
                    yield obs
                    return

                self.history.append({"role": "system", "content": f"Observation: {obs}"})
                yield f"**Observation:**\n{obs}"

            else:
                self.history.append({"role": "assistant", "content": response})
                break

    def _preview_and_pend(self, operation_name: str, func, args: tuple, kwargs: dict, preview_text: str) -> str:
        """Stores operation as pending and returns a confirmation request signal."""
        self._pending_operation = (func, args, kwargs)
        return f"__CONFIRM__:{operation_name}\n\n{preview_text}"

    def _execute_tool(self, action_str: str) -> str:
        """Parses tool call, generates preview for write ops, executes read ops directly."""
        try:
            tool_map = {
                "get_schema": self.tools.get_schema,
                "query_data": self.tools.query_data,
                "get_row_index": self.tools.get_row_index,
                "find_missing_rows": self.tools.find_missing_rows,
                "summarize_data": self.tools.summarize_data,
                "compute_and_query": self.tools.compute_and_query,
                "add_record": self.tools.add_record,
                "add_records": self.tools.add_records,
                "update_record": self.tools.update_record,
                "delete_record": self.tools.delete_record,
                "delete_missing_rows": self.tools.delete_missing_rows,
                "delete_records": self.tools.delete_records
            }

            cleaned = action_str.replace("Action:", "").strip()
            name, raw_args = cleaned.split("(", 1)
            name = name.strip()
            raw_args = raw_args.rstrip(")")

            # Write operations — show preview first
            if name in ("add_record", "add_records", "update_record", "delete_record", "delete_missing_rows", "delete_records"):

                if name == "add_records":
                    file_key, records = eval(f"({raw_args})", {"tool_map": tool_map})
                    df, _ = self.tools._load_data(file_key)
                    preview = (
                        f"**Operation:** Add {len(records)} record(s) to `{file_key}`\n\n"
                        f"**New records:**\n{pd.DataFrame(records).to_markdown(index=False)}\n\n"
                        f"**Current row count:** {len(df)} → **After:** {len(df) + len(records)}"
                    )
                    return self._preview_and_pend(
                        "Add Records",
                        tool_map[name], (), {"file_key": file_key, "records": records},
                        preview
                    )

                elif name == "add_record":
                    file_key, data = eval(f"({raw_args})", {"tool_map": tool_map})
                    df, _ = self.tools._load_data(file_key)
                    preview = (
                        f"**Operation:** Add new record to `{file_key}`\n\n"
                        f"**New record:**\n{pd.DataFrame([data]).to_markdown(index=False)}\n\n"
                        f"**Current row count:** {len(df)} → **After:** {len(df) + 1}"
                    )
                    return self._preview_and_pend(
                        "Add Record",
                        tool_map[name], (), {"file_key": file_key, "data": data},
                        preview
                    )

                elif name == "update_record":
                    file_key, row_index, updates = eval(f"({raw_args})", {"tool_map": tool_map})
                    df, _ = self.tools._load_data(file_key)
                    original = df.iloc[int(row_index)].to_dict()
                    preview = (
                        f"**Operation:** Update record at index `{row_index}` in `{file_key}`\n\n"
                        f"**Before:**\n{pd.DataFrame([original]).to_markdown(index=False)}\n\n"
                        f"**After:**\n{pd.DataFrame([{**original, **updates}]).to_markdown(index=False)}"
                    )
                    return self._preview_and_pend(
                        "Update Record",
                        tool_map[name], (), {"file_key": file_key, "row_index": row_index, "updates": updates},
                        preview
                    )

                elif name == "delete_record":
                    file_key, row_index = eval(f"({raw_args})", {"tool_map": tool_map})
                    df, _ = self.tools._load_data(file_key)
                    row = df.iloc[int(row_index)].to_dict()
                    preview = (
                        f"**Operation:** Delete record at index `{row_index}` from `{file_key}`\n\n"
                        f"**Record to be deleted:**\n{pd.DataFrame([row]).to_markdown(index=False)}\n\n"
                        f"**Current row count:** {len(df)} → **After:** {len(df) - 1}"
                    )
                    return self._preview_and_pend(
                        "Delete Record",
                        tool_map[name], (), {"file_key": file_key, "row_index": row_index},
                        preview
                    )

                elif name == "delete_missing_rows":
                    file_key = eval(f"({raw_args})", {"tool_map": tool_map})
                    if isinstance(file_key, tuple):
                        file_key = file_key[0]
                    df, _ = self.tools._load_data(file_key)
                    missing = df[df.isnull().any(axis=1)]
                    preview = (
                        f"**Operation:** Delete all rows with missing values from `{file_key}`\n\n"
                        f"**Rows to be deleted ({len(missing)}):**\n{missing.to_markdown(index=True)}\n\n"
                        f"**Current row count:** {len(df)} → **After:** {len(df) - len(missing)}"
                    )
                    return self._preview_and_pend(
                        "Delete Missing Rows",
                        tool_map[name], (), {"file_key": file_key},
                        preview
                    )
                elif name == "delete_records":
                    file_key, row_indices = eval(f"({raw_args})", {"tool_map": tool_map})
                    df, _ = self.tools._load_data(file_key)
                    valid_indices = [int(i) for i in row_indices if int(i) in df.index]
                    rows = df.loc[valid_indices]
                    preview = (
                        f"**Operation:** Delete {len(valid_indices)} records from `{file_key}`\n\n"
                        f"**Records to be deleted:**\n{rows.to_markdown(index=True)}\n\n"
                        f"**Current row count:** {len(df)} → **After:** {len(df) - len(valid_indices)}"
                    )
                    return self._preview_and_pend(
                        "Delete Records",
                        tool_map[name], (), {"file_key": file_key, "row_indices": row_indices},
                        preview
                    )
            return str(eval(f"tool_map['{name}']({raw_args})"))

        except Exception as e:
            return f"Execution Error: {str(e)}. Ensure tool name and arguments are correct."

    async def chat(self, user_input: str) -> AsyncGenerator[str, None]:
        self.history.append({"role": "user", "content": user_input})
        
        for _ in range(5):
            full_prompt = SYSTEM_PROMPT + "\n" + "\n".join(
                [f"{m['role']}: {m['content']}" for m in self.history]
            )
            
            response = self._call_llm(full_prompt)
            if not response:
                yield "No response from model."
                break

            lines = response.split("\n")
            truncated_lines = []
            for line in lines:
                truncated_lines.append(line)
                if line.strip().startswith("Action:") and "(" in line:
                    break
            response = "\n".join(truncated_lines)

            yield response

            if "Final Answer:" in response:
                self.history.append({"role": "assistant", "content": response})
                break

            elif "Action:" in response:
                action_line = next(
                    (line for line in response.split("\n") if line.strip().startswith("Action:") and "(" in line),
                    None
                )

                if not action_line or "none" in action_line.lower():
                    self.history.append({"role": "assistant", "content": response})
                    break

                observation = self._execute_tool(action_line)
                self.history.append({"role": "assistant", "content": response})

                # If it's a confirmation request, yield it and pause the loop
                if observation.startswith("__CONFIRM__:"):
                    yield observation
                    return  # Pause, app.py will call confirm_pending() or cancel_pending()

                self.history.append({"role": "system", "content": f"Observation: {observation}"})
                yield f"**Observation:**\n{observation}"

            else:
                self.history.append({"role": "assistant", "content": response})
                break